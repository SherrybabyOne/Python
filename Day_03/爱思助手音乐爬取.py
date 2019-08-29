
# [1]导入包
# 1.1网络访问
import requests;
# 1.2数据解析
from bs4 import BeautifulSoup
# 1.3自动化存储表格
from openpyxl import Workbook;
from openpyxl import load_workbook
# 1.4 导入os
import os;
# 苹果电脑
import ssl
ssl.create_default_context = ssl._create_unverified_context;
# 保存文件地址：
file = "2019年8月29日-爱思助手.xlsx";
# 模拟谷歌浏览器进行网络访问
# User-Agent: Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.132 Safari/537.36
# User-Agent: Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.102 Safari/537.36 Edge/18.18362
header = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.132 Safari/537.36"}
# 目标网址
mp3_url = "https://www.i4.cn/%s.html"
#https://www.i4.cn/ring_22_0_1.html
# [5]访问网络获取数据
def Get_data_url(url,ws):
    # 5.1 进行网址的拼接
    new_url = mp3_url%url;
    print(new_url)
    # 5.2开始网络访问
    request_url = requests.get(new_url,header);
    # 5.3判断是否访问成功
    print("网络请求状态码:",request_url.status_code);
    # 5.4获取数据
    if request_url.status_code == 200:
        # decode("utf-8") 添加数据的编码格式
        request_data = request_url.content.decode("utf-8");
        print("数据为;============================================")
        # print(request_data)
        # 本地数据存储 - 练习
        # 5.5 数据解析
        soup = BeautifulSoup(request_data,"html.parser");
        # 5.5.1 获取音乐容器
        div_mp3 = soup.find_all("div",attrs={"class":"list ring_list"});
        # 5.5.2获取每一条音乐的资料
        for i in range(0,len(div_mp3)):
            print("==========第%d首音乐信息=========="%i);
            # print(div_mp3[i]);
            div_item = div_mp3[i]
            # 01.音乐名称
            mp3_name = div_item.find("div",attrs={"class":"title"}).string;
            # 02.音乐下载量
            mp3_downcount = div_item.find("div", attrs={"class": "downcount"}).string;
            # 03.音乐时长
            mp3_longtime = div_item.find("div", attrs={"class": "longtime"}).string;
            # 04.音乐下载地址
            mp3_url_01 = div_item.find("div", attrs={"class": "btn audio_play"})["data-mp3"];
            print("音乐名称:",mp3_name)
            print("音乐下载量:", mp3_downcount)
            print("音乐时长:", mp3_longtime)
            print("音乐下载地址:", mp3_url_01)
            # 保存到表格中
            ws.append([mp3_name,mp3_downcount,mp3_longtime,mp3_url_01])
            pass
        pass
    else:
        print("网络访问失败")
        pass
    pass

# [2]程序的入口 "__"双下划线
if __name__ == '__main__':
    # 创建表格
    # [3]判断文件是否存在
    if os.path.exists(file):
        print("文件存在，进行文件读取")
        pass
    else:
        print("文件不存在,进行新建保存")
        wb = Workbook();
        wb.save(file)
        pass
    # 删除选项卡
    print("请选取要爬取的音乐:")
    sheet_list = ["1.月排行","2.总排行"];
    sheet_url_list = ["ring_22_0_1","ring_23_0_1"];
    wb = load_workbook(file);
    sheeet_list = wb.sheetnames;
    for i in range(0,len(sheet_list)):
        if sheet_list[i] not in sheeet_list:
            wb.create_sheet(sheet_list[i], i);
            pass
        print(sheet_list[i])
        pass
    # [3]控制台输入语句
    input_01 = int(input("请输入编号\n"))
    print(type(input_01));
    # 选择要插入数据的选项卡 ：1.月排行
    ws = wb[sheet_list[input_01-1]];
    # [4]开始请求网络访问数据
    Get_data_url(sheet_url_list[input_01-1],ws)
    # 数据保存
    wb.save(file)
    pass
# 拓展题- 不删除原有数据的情况下进行数据保存