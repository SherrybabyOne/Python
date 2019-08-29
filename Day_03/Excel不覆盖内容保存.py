
# [1]导包
from openpyxl import Workbook;
from openpyxl import load_workbook
import os
# [2]本地文件为
file = "2019年8月29日.xlsx";
# [3]判断文件是否存在
if os.path.exists(file):
    print("文件存在，进行文件读取")
    pass
else:
    print("文件不存在,进行新建保存")
    wb = Workbook();
    wb.save(file)
    pass
# [4]读取本地文件，进行数据修改
wb = load_workbook(file);
# [5]查询原有的选项卡
sheeet_list = wb.sheetnames;
print("一共有：%d个选项卡"%len(sheeet_list));
print(sheeet_list)
names = ["总排行","月排行","周排行"]
# [6]用于查看是否有对应的选项卡
for i in range(0,len(names)):
    if names[i] not in sheeet_list:
        wb.create_sheet(names[i],i);
        pass
    pass
# 修改一下数据
ws_01 = wb[names[0]];
ws_02 = wb[names[1]];
ws_03 = wb[names[2]];

# 新添加的数据
ws_01["A2"] = "7654321"

ws_02["B3"] = "1234567"

# 保存
wb.save(file)